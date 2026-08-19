import sys
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def apply_header_style(ws):
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 30


def create_cluster_summary(wb, keywords):
    """Create the Topic Cluster Summary tab as the first sheet in the workbook."""
    ws = wb.create_sheet("Topic Cluster Summary", 0)

    # Aggregate data by cluster
    clusters = {}
    for kw in keywords:
        cluster = kw.get("Topic Cluster", "Unknown")
        if cluster not in clusters:
            clusters[cluster] = {
                'count': 0,
                'us_volume': 0,
                'global_volume': 0,
                'kd_values': [],
                'cpc_values': [],
                'high': 0,
                'medium': 0,
                'low': 0,
            }
        clusters[cluster]['count'] += 1
        clusters[cluster]['us_volume'] += kw.get('Monthly Volume (US)', 0)
        clusters[cluster]['global_volume'] += kw.get('Global Volume', 0)
        clusters[cluster]['kd_values'].append(kw.get('Keyword Difficulty (KD)', 0))
        clusters[cluster]['cpc_values'].append(kw.get('CPC', 0))
        priority = str(kw.get('Priority', '')).lower()
        if priority == 'high':
            clusters[cluster]['high'] += 1
        elif priority == 'medium':
            clusters[cluster]['medium'] += 1
        else:
            clusters[cluster]['low'] += 1

    # Sort clusters by US volume descending
    sorted_clusters = sorted(clusters.items(), key=lambda x: x[1]['us_volume'], reverse=True)

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='2B579A')
    data_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', bold=True, size=11)
    total_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Topic Cluster Search Volume Summary'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 25

    # Subtitle row
    ws.merge_cells('A2:I2')
    num_clusters = len(sorted_clusters)
    num_keywords = sum(c['count'] for _, c in sorted_clusters)
    ws['A2'] = f'{num_keywords} Keywords | {num_clusters} Topic Clusters'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws.row_dimensions[2].height = 18

    # Table headers (row 4)
    headers = [
        'Topic Cluster', 'Keywords', 'Monthly Volume (US)', 'Global Volume',
        'Avg KD', 'Avg CPC ($)', 'High Priority', 'Medium Priority', 'Low Priority'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    # Data rows
    totals = {'count': 0, 'us_volume': 0, 'global_volume': 0, 'kd_all': [], 'cpc_all': [], 'high': 0, 'medium': 0, 'low': 0}

    for row_idx, (cluster_name, stats) in enumerate(sorted_clusters, 5):
        avg_kd = round(sum(stats['kd_values']) / len(stats['kd_values']), 1) if stats['kd_values'] else 0
        avg_cpc = round(sum(stats['cpc_values']) / len(stats['cpc_values']), 2) if stats['cpc_values'] else 0

        row_data = [
            cluster_name, stats['count'], stats['us_volume'], stats['global_volume'],
            avg_kd, avg_cpc, stats['high'], stats['medium'], stats['low']
        ]

        totals['count'] += stats['count']
        totals['us_volume'] += stats['us_volume']
        totals['global_volume'] += stats['global_volume']
        totals['kd_all'].extend(stats['kd_values'])
        totals['cpc_all'].extend(stats['cpc_values'])
        totals['high'] += stats['high']
        totals['medium'] += stats['medium']
        totals['low'] += stats['low']

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx in [3, 4]:
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.number_format = '$#,##0.00'

    # Totals row
    total_row = len(sorted_clusters) + 5
    total_avg_kd = round(sum(totals['kd_all']) / len(totals['kd_all']), 1) if totals['kd_all'] else 0
    total_avg_cpc = round(sum(totals['cpc_all']) / len(totals['cpc_all']), 2) if totals['cpc_all'] else 0
    total_data = [
        'TOTAL', totals['count'], totals['us_volume'], totals['global_volume'],
        total_avg_kd, total_avg_cpc, totals['high'], totals['medium'], totals['low']
    ]

    for col_idx, value in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_idx in [3, 4]:
            cell.number_format = '#,##0'
        elif col_idx == 6:
            cell.number_format = '$#,##0.00'

    # Volume share note
    note_row = total_row + 2
    ws.merge_cells(f'A{note_row}:I{note_row}')
    ws.cell(row=note_row, column=1, value='Volume Share by Cluster:').font = Font(name='Calibri', bold=True, size=11, color='2B579A')

    share_row = note_row + 1
    shares = []
    for cluster_name, stats in sorted_clusters:
        pct = round(stats['us_volume'] / totals['us_volume'] * 100, 1) if totals['us_volume'] > 0 else 0
        shares.append(f"{cluster_name}: {pct}%")
    ws.merge_cells(f'A{share_row}:I{share_row}')
    ws.cell(row=share_row, column=1, value=' | '.join(shares)).font = Font(name='Calibri', size=10, color='444444')

    # Column widths
    col_widths = [35, 10, 22, 18, 10, 12, 14, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def create_excel(json_data_path, output_path):
    with open(json_data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()

    # Tab 1: Priority Keyword Targets
    ws1 = wb.active
    ws1.title = "Priority Keyword Targets"

    keywords = data.get("keywords", [])
    if len(keywords) != 300:
        print(f"Warning: Expected 300 keywords, got {len(keywords)}.")

    df_keywords = pd.DataFrame(keywords)
    expected_cols = [
        "Keyword", "Topic Cluster", "Search Intent", "Monthly Volume (US)",
        "Global Volume", "Keyword Difficulty (KD)", "KD Level", "CPC",
        "Parent Topic", "Recommended Content Type", "Priority"
    ]

    # Ensure all columns exist
    for col in expected_cols:
        if col not in df_keywords.columns:
            df_keywords[col] = ""

    df_keywords = df_keywords[expected_cols]

    for r in dataframe_to_rows(df_keywords, index=False, header=True):
        ws1.append(r)

    apply_header_style(ws1)

    # Apply color coding for Priority and KD Level
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    priority_col_idx = expected_cols.index("Priority") + 1
    kd_level_col_idx = expected_cols.index("KD Level") + 1

    for row in range(2, ws1.max_row + 1):
        # Priority colors
        priority_cell = ws1.cell(row=row, column=priority_col_idx)
        val = str(priority_cell.value).lower()
        if val == 'high':
            priority_cell.fill = green_fill
        elif val == 'medium':
            priority_cell.fill = yellow_fill
        elif val == 'low':
            priority_cell.fill = red_fill

        # KD Level formatting
        kd_cell = ws1.cell(row=row, column=kd_level_col_idx)
        kd_val = str(kd_cell.value).lower()
        if kd_val == 'easy':
            kd_cell.fill = green_fill
        elif kd_val == 'medium':
            kd_cell.fill = yellow_fill
        elif kd_val == 'hard':
            kd_cell.fill = red_fill

    # Adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws1.column_dimensions[column].width = min(adjusted_width, 40)

    # Tab 2: Competitor Keyword Gaps & Landscape
    ws2 = wb.create_sheet(title="Competitor Gaps & Landscape")

    row_idx = 1
    bold_font = Font(bold=True, size=12)

    # Section 1
    ws2.cell(row=row_idx, column=1, value="Section 1: Traffic Share by Domain").font = bold_font
    row_idx += 2

    competitors = data.get("competitors", [])
    df_comp = pd.DataFrame(competitors)
    comp_cols = ["Domain", "Traffic Share %", "Traffic Value", "DR", "Observations"]
    for col in comp_cols:
        if col not in df_comp.columns:
            df_comp[col] = ""
    df_comp = df_comp[comp_cols]

    for r in dataframe_to_rows(df_comp, index=False, header=True):
        for c_idx, val in enumerate(r, 1):
            cell = ws2.cell(row=row_idx, column=c_idx, value=val)
            if row_idx == 3:  # Header
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True)
        row_idx += 1

    row_idx += 2

    # Section 2
    ws2.cell(row=row_idx, column=1, value="Section 2: Keyword Gap Opportunities").font = bold_font
    row_idx += 2

    gaps = data.get("gaps", [])
    df_gaps = pd.DataFrame(gaps)
    gap_cols = ["Competitor Domain", "Keyword Gap", "Search Volume", "Recommended Action"]
    for col in gap_cols:
        if col not in df_gaps.columns:
            df_gaps[col] = ""
    df_gaps = df_gaps[gap_cols]

    start_row = row_idx
    for r in dataframe_to_rows(df_gaps, index=False, header=True):
        for c_idx, val in enumerate(r, 1):
            cell = ws2.cell(row=row_idx, column=c_idx, value=val)
            if row_idx == start_row:  # Header
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True)
        row_idx += 1

    row_idx += 2

    # Section 3
    ws2.cell(row=row_idx, column=1, value="Section 3: Strategic Insights & Recommendations").font = bold_font
    row_idx += 2

    insights = data.get("insights", {})
    ws2.cell(row=row_idx, column=1, value="Key Highlights:").font = Font(bold=True)
    row_idx += 1
    ws2.cell(row=row_idx, column=1, value=f"Total Addressable Market: {insights.get('tam', '')}")
    row_idx += 1
    ws2.cell(row=row_idx, column=1, value=f"Best Immediate Opportunities: {insights.get('opportunities', '')}")
    row_idx += 2

    ws2.cell(row=row_idx, column=1, value="5 Key Findings:").font = Font(bold=True)
    row_idx += 1
    for finding in insights.get("findings", []):
        ws2.cell(row=row_idx, column=1, value=f"• {finding}")
        row_idx += 1

    row_idx += 1
    ws2.cell(row=row_idx, column=1, value="5 Prioritized Recommendations:").font = Font(bold=True)
    row_idx += 1
    for i, rec in enumerate(insights.get("recommendations", []), 1):
        ws2.cell(row=row_idx, column=1, value=f"{i}. {rec}")
        row_idx += 1

    # Adjust widths for Tab 2
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 40

    # Create Topic Cluster Summary tab (inserted as first sheet)
    create_cluster_summary(wb, keywords)

    wb.save(output_path)
    print(f"Successfully generated Excel workbook at {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_keyword_excel.py <input_json_path> <output_excel_path>")
        print("Note: The JSON file should contain 'keywords', 'competitors', 'gaps', and 'insights'.")
        sys.exit(1)

    input_json = sys.argv[1]
    output_excel = sys.argv[2]

    create_excel(input_json, output_excel)
